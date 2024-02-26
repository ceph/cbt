#!/usr/bin/env python3
import yaml
import argparse
import os
import time
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
import sys

''' directory structure:
    autobench.{date}/rep-{repeat_id}/test-{test_id}
    test-{test_id} is the output directory from crimson_bench_tool.

    graphic directory:
    autobench.{date}.graphic

    cross analyse graphic directory:
    autobench.{date1}.autobench.{date2}.graphic
'''

no_value_attributes= ['crimson', 'output_horizontal', 'perf', \
                      'perf_record', 'iostat', 'emon']

# transfer to crimson_bench_tool param
def trans(param):
    res = f"--{param}"
    res = res.replace('_', '-')
    return res

def prefix_match(prefix, target):
    prefix = prefix.lower()
    target = target.lower()
    if target[:len(prefix)] == prefix:
        return True
    else:
        return False

def get_version_and_commitID():
    month_dic={
        "Jan":"01", "Feb":"02", "Mar":"03", "Apr":"04",
        "May":"05", "Jun":"06", "Jul":"07", "Aug":"08",
        "Sep":"09", "Oct":"10", "Nov":"11", "Dec":"12",
    }
    gitlog = os.popen("git log ..")
    line = gitlog.readline()
    commitID = None
    version = None
    while line:
        ll = line.split()
        if ll[0] == "commit" :
            commitID = ll[1][:8]
        if ll[0] == "Date:":
            version = ll[5] + month_dic[ll[2]] + ll[3]
            break
        line = gitlog.readline()
    return version, commitID

def root_protect(dev):
    protect_dir = '/'
    tgt_par = dev.split('/')[-1]
    lsblk = os.popen("lsblk")
    disk = None
    line = lsblk.readline()
    while line:
        ll = line.split()
        if ll[0][0:2] == '├─' or ll[0][0:2] == '└─':
            par = ll[0][2:]
        if ll[0][0:2] != '├─' and ll[0][0:2] != '└─':
            disk = ll[0]
        if ll[-1] == protect_dir:
            if disk == tgt_par or par == tgt_par:
                raise Exception("Be awake! You should not write "\
                                "to the disk/partition that root exist!")
        line = lsblk.readline()

def test_device(sys_info_path, dev):
    root_protect(dev)
    configs = [{'mode':'randwrite', 'bs':'4K', 'time':'30'},
               {'mode':'randread', 'bs':'4K', 'time':'30'},
               {'mode':'write', 'bs':'1M', 'time':'30'},
               {'mode':'read', 'bs':'1M', 'time':'30'}]
    for config in configs:
        bs = config['bs']
        mode = config['mode']
        time = config['time']
        print(f'doing {time}s device {dev} {mode} test...')
        cmd = f'fio --filename={dev} '\
            f'--numjobs=8 --iodepth=6 --ioengine=libaio --direct=1 --verify=0 '\
            f'--bs={bs}  --rw={mode} --group_reporting=1 '\
            f'--runtime={time} --time_based --name=device_test'
        ret = os.popen(cmd)
        with open(sys_info_path, "a") as f:
            f.write(f'disk {mode} test result:\n')
            f.write(ret.read())
            f.write('\n')
        f.close()

def record_system_info(root, configs):
    path = f"{root}/sys_info.txt"
    os.system(f"touch {path}")

    os.system(f"echo \"0.Ceph Source Code Info\">> {path}")
    version, commitID = get_version_and_commitID()
    os.system(f"echo \"version: {version}\" >> {path}")
    os.system(f"echo \"commit ID: {commitID}\" >> {path}")
    os.system(f"echo >> {path}")

    os.system(f"echo \"1.Linux Release\">> {path}")
    os.system(f"lsb_release -a >> {path}")
    os.system(f"echo >> {path}")

    os.system(f"echo \"2.Linux Kernel\">> {path}")
    os.system(f"uname -r  >> {path}")
    os.system(f"echo >> {path}")

    os.system(f"echo \"3.CPU\" >> {path}")
    os.system(f"lscpu >> {path}")
    os.system(f"echo >> {path}")

    os.system(f"echo \"4.Disk\" >> {path}")
    os.system(f"lsblk -O --raw -a >> {path}")
    os.system(f"echo >> {path}")
    os.system(f"sudo fdisk -l >> {path}")
    os.system(f"echo >> {path}")
    # will only show device detail configed in config yaml
    devs = set()
    for config in configs:
        if 'dev' in config:
            devs.add(config['dev'])
    for dev_id, dev in enumerate(devs):
        os.system(f"echo \"4.{dev_id+1}.1 Disk: {dev}\" >> {path}")
        if dev[5:8] == 'nvm':
            os.system(f"sudo nvme id-ctrl {dev} >> {path}")
        else:
            os.system(f"sudo hdparm -I {dev} >> {path}")
        os.system(f"echo >> {path}")
        if not no_disk_test:
            os.system(f"echo \"4.{dev_id+1}.2 Disk Test\" >> {path}")
            print('start basic disk tests... you can add --no-disk-test to skip this.')
            test_device(path, dev)
            os.system(f"echo >> {path}")
        else:
            print('will not do basic disk test.')

    os.system(f"echo \"5.Memory\" >> {path}")
    os.system(f"lsmem >> {path}")
    os.system(f"echo >> {path}")
    os.system(f"sudo dmidecode -t memory >> {path}")
    os.system(f"echo >> {path}")

    os.system(f"echo \"6.Board\" >> {path}")
    os.system(f"sudo dmidecode|grep -A16 \"System Information$\" >> {path}")
    os.system(f"echo >> {path}")

    os.system(f"echo \"7.PCIe\" >> {path}")
    os.system(f"lspci -vv >> {path}")
    os.system(f"echo >> {path}")

# e.g. config_file='./config.yaml', x='osd_cores', comp=[1, 2, 3]
def read_config(config_file, x = None, comp = None):
    config_file_pwd = f"{config_file}"
    f = open(config_file_pwd, 'r')
    _config = yaml.safe_load_all(f.read())
    f.close()

    configs = list()
    for config in _config:
        configs.append(config)

        for key in config:
            # key check
            if key in no_value_attributes and config[key] != True:
                raise Exception("Error: no value attributes should only be True")
            if key == 'alias':
                alias = str(config[key])
                if len(alias.split()) != 1:
                    raise Exception("Error: alias should not include space")

    if x:
        for config in configs:
            if args.x not in config:
                raise Exception("Error: x param should exist in config yaml file")
    if comp:
        for index in comp:
            if index > len(configs) or index <= 0:
                raise Exception(f"Error: comp index error. Shoud between" \
                                f" 1-{len(configs)}")
    return configs

def do_bench(config_file, configs, repeat, build, output):
    # all files' root for this auto bench test(bench result).
    # will be stored at where this tool is.
    root = f"{current_path}"
    if output:
        root = f"{root}/{output}"
    else:
        with os.popen("date +%Y%m%d.%H%M%S") as date:
            line = date.readline()
            res = line.split()[0]
            root = f"{root}/autobench.{res}"
    delete_and_create_at(root)
    print('=======================')
    config_path = args.run if args.run else args.bench
    print(f'Using config file path: {config_path}')
    print(f'Using ceph build path: {build}')
    print(f'Auto bench result path: {root}')
    print('=======================')
    record_system_info(root, configs)
    os.system(f"cp {current_path}/{config_file} {root}/config.yaml")

    # do bench
    for repeat_id in range(repeat):
        repeat_path = f"{root}/rep-{repeat_id}"
        os.makedirs(repeat_path)
        for test_id, test_config in enumerate(configs):
            command = f"{tool_path}/crimson_bench_tool.py"
            alias = None
            for key in test_config:
                if key in no_value_attributes:
                    command += f" {trans(key)}"
                elif key == 'alias':
                    alias = test_config[key]
                else:
                    command += f" {trans(key)} {test_config[key]}"
            test_path_prefix = f"{repeat_path}/test-{test_id}"
            if alias:
                test_path_prefix += f"_{alias}"
            command += f" --log {test_path_prefix}"
            command += f" --build {build}"
            print(command)
            print(f'testing... repeat: {repeat_id+1}, test: {test_id+1}')
            os.system(command)
            if (test_id != len(configs) - 1) or (repeat_id != repeat - 1):
                print(f'test gap time: {gap}s')
                time.sleep(gap)
    read_results(root, len(configs))
    return root

def read_results(root, test_num):
    root_files = os.listdir(root)
    failure_log = 'failure_log.txt'
    failure_path = f"{root}/{failure_log}"
    first = False
    if failure_log not in root_files:
        first = True
        os.system(f"touch {failure_path}")

    # read root directory to get results
    results = dict()
    # results - repeat_results - test_results - (tester_id, all results)
    repeat_num = len(os.listdir(root)) - 3 # skip config.yaml, failure_log.txt, sys_info.txt
    for repeat_id in range(repeat_num):
        repeat_path = f"{root}/rep-{repeat_id}"
        repeat_results = dict()

        tests = os.listdir(repeat_path)
        for test_id in range(test_num):
            tgt_test_dir_prefix = f'test-{test_id}'
            test_dir = None
            # check if dir exits
            for dir in tests:
                if dir.split('_')[0] == tgt_test_dir_prefix:
                    test_dir = dir
                    break
            if test_dir:
                test_path = f"{repeat_path}/{test_dir}"
                test_results = dict()
                files = os.listdir(test_path)
                if '__failed__' in files:
                    repeat_results[test_id] = False
                    if first:
                        os.system(f"echo \"[TEST FAILED] in rep-{repeat_id},"\
                                f" test-{test_id}\" >> {failure_path}")
                else:
                    result_path = f"{test_path}/result.csv"
                    if os.path.exists(result_path):
                        res = pd.read_csv(result_path)
                        for case in range(len(res)):
                            test_results[case] = res.iloc[case]
                        repeat_results[test_id] = test_results
                    else:
                        repeat_results[test_id] = False
                        if first:
                            os.system(f"echo \"[INPUT CONFIG ERROR or UNKOWN ERROR] in "\
                                    f"rep-{repeat_id}, test-{test_id}, please run the "\
                                    f"{test_id}th test using the {test_id}th group of "\
                                    f"param in your config file by crimson_bench_tool.py "\
                                    f"to get the error information.\" >> {failure_path}")
            else:
                repeat_results[test_id] = False
                if first:
                    os.system(f"echo \"[INPUT CONFIG ERROR or UNKOWN ERROR] in "\
                            f"rep-{repeat_id}, test-{test_id}, please run the "\
                            f"{test_id}th test using the {test_id}th group of "\
                            f"param in your config file by crimson_bench_tool.py "\
                            f"to get the error information.\" >> {failure_path}")

        results[repeat_id] = repeat_results
    return results

def print_results(results):
    for repeat_id in results:
        for test_id in results[repeat_id]:
            print(f"repeat_id:{repeat_id}, test_id:{test_id}")
            print(results[repeat_id][test_id])

def delete_and_create_at(path):
    if os.path.exists(path):
        os.system(f"sudo rm -rf {path}")
    os.makedirs(path)

def adjust_results(results, y):
    '''
        transfer to [[[r1t1c1y, r2t1c1y, r3t1c1y], [r1t1c2y, r2t1c2y, r3t1c3y], ...], => pics1
                     [[r1t2c1y, r2t2c1y, r3t2c1y], [r1t2c2y, r2t2c2y, r3t2c3y], ...],  => pics2
                     ...]
        r1t1c1y means repeat 1, test 1, case 1, y
        t - c - r
        one pics for one test case, correspoding to one config block in yaml file
        the most inner list represent multiple repeat target y for one tester(case) in one test
    '''
    test_num = len(results[0])
    all_test_res = list()
    for test_id in range(test_num):
        case_num = 0
        for repeat_id in results:
            if results[repeat_id][test_id]:
                _case_num = len(results[repeat_id][test_id])
                if case_num != 0 and _case_num != case_num:
                    raise Exception("Error: cases num changed\
                                    between different repeat for one same test")
                case_num = _case_num
        if case_num == 0:
            # all repeat of this test failed
            all_test_res.append([])
            continue

        test_res = list()
        for case_id in range(case_num):
            case_res = list()
            for repeat_id in results:
                if results[repeat_id][test_id] != False:
                    match_count = 0
                    case = results[repeat_id][test_id][case_id]
                    for y_f in case.keys().to_list():
                        if prefix_match(y, y_f):
                            y_res = results[repeat_id][test_id][case_id][y_f]
                            match_count += 1
                    if match_count > 1:
                        raise Exception(f"Error: {y} matches multiple y indexes"\
                                        f", please input full y index name")
                    if match_count == 0:
                        raise Exception(f"Error: {y} didn't match any y index")
                    case_res.append(y_res)
                else:
                    pass
            test_res.append(case_res)
        all_test_res.append(test_res)
    return all_test_res

def draw(m_analysed_results, m_configs, x, y, res_path, m_comp, alias, m_repnums):
    res_path = f'{current_path}/{res_path}'
    delete_and_create_at(res_path)
    color_set = ['b', 'g', 'r', 'c', 'm', 'y', 'k']
    color_set_p = 0
    for auto_bench_id, analysed_results in enumerate(m_analysed_results):
        configs = m_configs[auto_bench_id]
        output_auto_bench_name = None
        if alias:
            output_auto_bench_name = f'{alias[auto_bench_id]}_'
        else:
            if len(m_analysed_results) > 1:
                output_auto_bench_name = f"bench_{auto_bench_id}_"
            else:
                output_auto_bench_name = ''
        comp = None
        if m_comp:
            comp = m_comp[auto_bench_id]
        repnum_title = f'repeat:{m_repnums[0]}'
        for rep_id in range(1, len(m_repnums)):
            repnum_title += f',{m_repnums[rep_id]}'
        for test_id, test in enumerate(analysed_results):
            if test == []:
                print(f'test {test_id} failed')
                continue
            if comp and test_id+1 not in comp:
                continue
            x_value = configs[test_id][x]
            if type(x_value) == int:
                x_data = [x_value]
            else:
                x_data = x_value.split()
            test_alias = None
            if 'alias' in configs[test_id]:
                test_alias = configs[test_id]['alias']
            else:
                test_alias = f"test_{test_id}"
            y_data = test
            y_data_mean = list()
            for items in y_data:
                y_data_mean.append(sum(items)/len(items))

            df = pd.DataFrame({f'{x}':[], f'{y}':[]})
            for x_id, x_content in enumerate(x_data):
                for y_content in y_data[x_id]:
                    df.loc[len(df.index)] = {f'{x}': x_content, f'{y}' : y_content}

            plt.title(f"{x}-{y} ({repnum_title})".lower())
            plt.xlabel(f"{x}")
            plt.ylabel(f"{y}")
            color = color_set[color_set_p]
            color_set_p += 1
            if color_set_p >= len(color_set):
                color_set_p = 0
            plt.plot(f'{x}', f'{y}', data=df, linestyle='none',\
                    marker='o', label=f'{output_auto_bench_name}{test_alias}', color=color)
            plt.plot(x_data, y_data_mean, linestyle='-', \
                     label=f'{output_auto_bench_name}{test_alias} mean', color=color)
            plt.grid(True, color='gray', linestyle='--')
            plt.legend(loc=2)
            plt.rc('legend', fontsize='x-small')
            # TODO: additional information to graphics
            if not comp:
                if start_from_zero:
                    plt.ylim(ymin=0)
                plt.savefig(f"{res_path}/{output_auto_bench_name}"\
                            f"{test_alias}_x-{x}_y-{y}.png".lower(), dpi=500)
                plt.close()

            # raw data to csv
            df.to_csv(f"{res_path}/{output_auto_bench_name}"\
                      f"{test_alias}_x-{x}_y-{y}.csv".lower())
            # average to csv
            df_avg = pd.DataFrame({f'{x}':[], f'{y}_avg':[]})
            for x_id, x_content in enumerate(x_data):
                df_avg.loc[len(df_avg.index)] = \
                    {f"{x}": x_content, f'{y}_avg' : y_data_mean[x_id]}
            df_avg.to_csv(f"{res_path}/{output_auto_bench_name}"\
                          f"{test_alias}_x-{x}_y-{y}_avg.csv".lower())

    if m_comp:
        if start_from_zero:
            plt.ylim(ymin=0)
        plt.savefig(f'{res_path}/x-{x}_y-{y}.png'.lower(), dpi=500)
        plt.close()

def process_emon(roots, m_configs, res_path, m_comp, alias, m_repnums):
    wb = openpyxl.Workbook()
    # find the common repeats in all roots
    repeat_num = sys.maxsize
    for rep in m_repnums:
        if rep < repeat_num:
            repeat_num = rep
    # generate sheet for each repeat with the results of
    # all cases in all tests in all auto bench roots.
    for rep_id in range(repeat_num):
        wbs = wb.create_sheet(f'rep-{rep_id}')
        test_dic = dict()
        for auto_bench_id, root in enumerate(roots):
            comp = m_comp[auto_bench_id] if m_comp else None
            configs = m_configs[auto_bench_id]
            col_auto_bench_name = None
            if alias:
                col_auto_bench_name = f'{alias[auto_bench_id]}-'
            else:
                if len(roots) > 1:
                    col_auto_bench_name = f"bench:{auto_bench_id}-"
                else:
                    col_auto_bench_name = ''

            rep_path = f"{root}/rep-{rep_id}"
            for test_dir in os.listdir(rep_path):
                test_prefix = test_dir.split('_')[0].split('-')
                if test_prefix[0] != 'test':
                    continue
                test_path = f'{rep_path}/{test_dir}'
                test_id = int(test_prefix[1])
                if comp and test_id+1 not in comp:
                    continue
                # a tester also means a case
                tester_dirs = sorted(os.listdir(test_path))
                for tester_dir in tester_dirs:
                    tester_prefix = tester_dir.split('.')[0]
                    # check if the length of first section is 1
                    if len(tester_prefix) != 1:
                        continue
                    tester_id = int(tester_prefix)
                    test_alias = None
                    if 'alias' in configs[test_id]:
                        test_alias = configs[test_id]['alias']
                    else:
                        test_alias = f"test:{test_id}"
                    col_name = f'{col_auto_bench_name}{test_alias}-case:{tester_id}'
                    tgt_emon_file = f'{test_path}/{tester_dir}/summary.xlsx'
                    if os.path.exists(tgt_emon_file):
                        # extract information from target emon file
                        emon_file = openpyxl.load_workbook(tgt_emon_file)
                        sheet = emon_file['socket view']
                        tester_dic = dict()
                        for row in range(1, sheet.max_row + 1):
                            name = sheet.cell(row=row, column=1).value
                            value = sheet.cell(row=row, column=2).value # use socket 0 value
                            tester_dic[name] = value
                        test_dic[col_name] = tester_dic
                        emon_file.close()

        # generate chossen name
        chossen_name = list()
        if emon_target:
            f_emon_target = open(emon_target, "r")
            line = f_emon_target.readline()
            while line:
                chossen_name.append(line.rstrip('\n'))
                line = f_emon_target.readline()
            f_emon_target.close()
        else:
            for col_name in test_dic:
                tester_dict = test_dic[col_name]
                for name in tester_dict:
                    if name not in chossen_name:
                        chossen_name.append(name)
        # fisrt row
        col_index = 'A'
        wbs['A1'] = 'Name'
        for col_name in test_dic:
            col_index = chr(ord(col_index)+1)
            wbs[f'{col_index}1'] = col_name
        # next rows
        for id, name in enumerate(chossen_name):
            row_index = id + 2
            col_index = 'A'
            wbs[f'A{row_index}'] = name
            for col_name in test_dic:
                col_index = chr(ord(col_index)+1)
                tester_dic = test_dic[col_name]
                value = None
                if name in tester_dic:
                    value = tester_dic[name]
                else:
                    value = 'not exist'
                wbs[f'{col_index}{row_index}'] = value
    del wb['Sheet']
    wb.save(res_path)

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='')
    parser.add_argument('--run',
                        type=str,
                        default=None,
                        help= "do bench, analyse results and draw pictures, require --x, --y. "\
                            "Input bench config yaml file path. "\
                            "You can add alias label for every test, which will show in graphic.")
    parser.add_argument('--bench',
                        type=str,
                        default=None,
                        help= "only do bench. Input bench config yaml file path. "\
                            "You can add alias label for every test, which will show in graphic.")
    parser.add_argument('--ana',
                        nargs='+',
                        type=str,
                        default=None,
                        help= "input the root directory storing the auto bench results, to"\
                            " adjust results and draw pictures. require --x, --y."\
                            " You can input mulitiple root bench results directory,"\
                            " and input --comp to do cross auto bench analyse")
    parser.add_argument('--comp',
                        nargs='+',
                        type=str,
                        default=None,
                        help= "will merge target index tests into one graphics. The "\
                            "index corresponds to the order in which the configuration "\
                            "appears in the config file."\
                            "Index in different auto bench results shuold be divided by"\
                            " space, index in same auto bench results should be divided by ,")
    parser.add_argument('--alias',
                        nargs='+',
                        type=str,
                        default=None,
                        help= "alias for each auto bench results correspoding to --ana or --emon"\
                            "This alias will show in output graphics/excel")

    parser.add_argument('--build',
                        type=str,
                        default='.',
                        help='build directory of ceph. Default to be .')
    parser.add_argument('--output',
                        type=str,
                        default=None,
                        help='bench results directory when --bench/--run(default to be autobench.date) or '\
                            'graphic results directory when --ana(default to be autobench dir name.graphic) or '\
                            'emon result file when --emon(default to be autobench dir name.emon)')
    parser.add_argument('--no-disk-test',
                        action='store_true',
                        help= "will not do basic device speed tests(which will cost you at"\
                            " least 2 minutes).")

    parser.add_argument('--repeat',
                        type=int,
                        default=1,
                        help="repeat time for every tests, default to be 1")
    parser.add_argument('--gap',
                        type=int,
                        default=1,
                        help="time gap between each test(run crimson_bench_tool once)")
    parser.add_argument('--x',
                        type=str,
                        help="x axis of result graphics, the main variable in the target"\
                            " graphics to draw x-y graphics. required when --ana or --run."\
                            " x can be osd_cores, client, thread, osd_op_num_shards, etc. all the"\
                            " parameters in the crimson bench tool can be x.")
    parser.add_argument('--y',
                        nargs='+',
                        type=str,
                        default=["IOPS"],
                        help="the label name of y asix of the result graphics, IOPS by default")
    parser.add_argument('--no-start-from-zero',
                        action='store_true',
                        help="y axis will not start from zero")

    parser.add_argument('--emon',
                        nargs='+',
                        type=str,
                        default=None,
                        help= "input the root directory storing the auto bench results, to"\
                            " intergrate and adjust all emon edp results into one document."\
                            " and input --comp to deicde using which groups of test results."\
                            " --emon also support cross auto bench analyse like --ana")
    parser.add_argument('--emon-target',
                        type=str,
                        default=None,
                        help= "input the path of file that record the emon results name you"\
                            " want. Each line of that file should be one name. None by default,"\
                            " which means all names will be used.")

    args = parser.parse_args()
    gap = args.gap
    res_path_suffix = 'graphic'
    no_disk_test = False
    start_from_zero = True
    emon_target = args.emon_target
    if args.no_disk_test:
        no_disk_test = True
    if args.no_start_from_zero:
        start_from_zero = False
    tool_path = (os.path.dirname(os.path.realpath(__file__)))
    current_path = os.getcwd()

    _run = 1 if args.run else 0
    _bench = 1 if args.bench else 0
    _ana = 1 if args.ana else 0
    _emon = 1 if args.emon else 0

    if _run + _bench + _ana + _emon != 1:
        raise Exception("Error: should run in one of run/bench/ana/emon")
    if args.run:
        if not args.x:
            raise Exception("Error: should input --x to run")
        if args.comp and len(args.comp) != 1:
            raise Exception("Error: should only do one auto bench when --run")
        comp = None
        if args.comp:
            comp = list()
            _comp = args.comp[0].split(',')
            for index in _comp:
                comp.append(int(index))
        configs = read_config(args.run, x=args.x, comp=comp)
        root = do_bench(args.run, configs, args.repeat, args.build, args.output)
        results = read_results(root, len(configs))
        print(root)
        # the path of graphic result will be the same as the bench result
        res_path = f"{root}.{res_path_suffix}"
        delete_and_create_at(res_path)
        for y in args.y:
            analysed_results = adjust_results(results, y)
            draw([analysed_results], [configs], args.x, y, res_path, \
                 [comp], args.alias)

    if args.bench:
        configs = read_config(args.bench)
        root = do_bench(args.bench, configs, args.repeat, args.build, args.output)
        print(root)

    if args.ana:
        roots = list()
        for root in args.ana:
            if root[-1] == '/':
                roots.append(root[:-1])
            else:
                roots.append(root)
        if not args.x:
            raise Exception("Error: should input --x to analyse")
        if args.comp and len(args.ana) != len(args.comp):
            raise Exception("Error: len of --comp shuold match the len of --ana")
        if args.alias and len(args.ana) != len(args.alias):
            raise Exception("Error: len of --alias shuold match the len of --ana")
        if len(args.ana) > 1 and not args.comp:
            raise Exception("Error: must use --comp when len of --ana > 1, which means"\
                            " you must use comp mode if you want to anaylse multiple auto"\
                            " bench results")
        m_configs = list()
        m_results = list()
        m_repnums = list()
        # comp process example: ["1,2,3", "1,2,3"] -> [[1,2,3],[1,2,3]]
        # structure like this is for the needs of cross bench results analyse.
        m_comp = None
        if args.comp:
            m_comp = list()
        res_path = ''
        for root_index, root in enumerate(roots):
            comp = None
            if args.comp:
                comp = list()
                _comp = args.comp[root_index].split(',')
                for index in _comp:
                    comp.append(int(index))
                m_comp.append(comp)

            configs = read_config(f"{current_path}/{root}/config.yaml", x=args.x, comp=comp)
            results = read_results(f"{current_path}/{root}", len(configs))

            res_path += f"{root}."
            m_configs.append(configs)
            m_results.append(results)
            # skip config.yaml, failure_log.txt, sys_info.txt
            repeat_num = len(os.listdir(root)) - 3
            m_repnums.append(repeat_num)
        # the path of graphic result will be the same as the bench result
        res_path += f"{res_path_suffix}"
        if args.output:
            res_path = args.output

        for y in args.y:
            m_analysed_results = list()
            for results in m_results:
                analysed_results = adjust_results(results, y)
                m_analysed_results.append(analysed_results)
            draw(m_analysed_results, m_configs, args.x, y, res_path, \
                 m_comp, args.alias, m_repnums)

        # copy sys_info.txt, config.yaml to graphic result directory
        for root_index, root in enumerate(roots):
            config_path = f"{current_path}/{root}/config.yaml"
            sys_info_path = f"{current_path}/{root}/sys_info.txt"
            tgt_config_path = f'{root_index}.config.yaml' if len(roots) != 1 else 'config.yaml'
            tgt_sys_info_path = f'{root_index}.sys_info.txt' if len(roots) != 1 else 'sys_info.txt'
            tgt_config_path = f'{current_path}/{res_path}/{tgt_config_path}'
            tgt_sys_info_path = f'{current_path}/{res_path}/{tgt_sys_info_path}'
            os.system(f'cp {config_path} {tgt_config_path}')
            os.system(f'cp {sys_info_path} {tgt_sys_info_path}')

    if args.emon:
        roots = list()
        for root in args.emon:
            if root[-1] == '/':
                roots.append(root[:-1])
            else:
                roots.append(root)
        if args.comp and len(args.emon) != len(args.comp):
            raise Exception("Error: len of --comp shuold match the len of --emon")
        if args.alias and len(args.emon) != len(args.alias):
            raise Exception("Error: len of --alias shuold match the len of --emon")
        if len(args.emon) > 1 and not args.comp:
            raise Exception("Error: must use --comp when len of --emon > 1, which means"\
                            " you must use comp mode if you want to anaylse multiple auto"\
                            " bench results")

        m_configs = list()
        m_repnums = list()
        m_comp = list() if args.comp else None
        res_path = ''
        for root_index, root in enumerate(roots):
            comp = None
            if args.comp:
                comp = list()
                _comp = args.comp[root_index].split(',')
                for index in _comp:
                    comp.append(int(index))
                m_comp.append(comp)

            configs = read_config(f"{current_path}/{root}/config.yaml", x=None, comp=comp)
            res_path += f"{root}."
            m_configs.append(configs)
            # skip config.yaml, failure_log.txt, sys_info.txts
            repeat_num = len(os.listdir(root)) - 3
            m_repnums.append(repeat_num)
        res_path += 'emon'
        if args.output:
            res_path = args.output
        res_path += '.xlsx'

        process_emon(roots, m_configs, res_path, m_comp, args.alias, m_repnums)
